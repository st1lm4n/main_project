import csv
import json
from typing import Any, Dict, List

from openpyxl import load_workbook


def read_json(file_path: str) -> List[Dict[str, Any]]:
    """Читает JSON-файл с транзакциями."""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def read_csv(file_path: str) -> List[Dict[str, Any]]:
    """Читает CSV-файл с транзакциями."""
    with open(file_path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def read_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """Читает XLSX-файл с транзакциями."""
    wb = load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    return [{headers[i]: cell.value for i, cell in enumerate(row)} for row in sheet.iter_rows(min_row=2)]


# print(read_csv('transactions.csv'))
# read_json('../data/operations.json')
# print(read_xlsx('..//data/transactions_excel.xlsx'))
